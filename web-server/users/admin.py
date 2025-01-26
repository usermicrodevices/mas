import sys, logging, xlsxwriter
from io import BytesIO
from datetime import datetime, timedelta

from django.contrib import admin, messages

from django import forms
from django.urls import reverse
from django.shortcuts import render
from django.db.models import Q, Value, CharField
from django.utils import timezone
from django.utils.html import format_html, format_html_join
from django.utils.translation import gettext as _
from django.http import StreamingHttpResponse, FileResponse

from django.contrib.auth.admin import UserAdmin as BaseUserAdmin
from django.contrib.auth.models import Permission
from django.contrib.contenttypes.models import ContentType

from django.conf import settings

from transliterate import slugify
from openpyxl import load_workbook

from .models import Role, RoleModel, RoleField, User
from django.apps import apps as django_apps


admin.site.subtitle = _('Users')


def get_model(app_model):
	app_name, model_name = app_model.split('.')
	return django_apps.get_app_config(app_name).get_model(model_name)


class UploadFileForm(forms.Form):
	_selected_action = forms.CharField(widget=forms.MultipleHiddenInput)
	file = forms.FileField(widget=forms.ClearableFileInput(attrs={'allow_multiple_selected': True}))


class CustomBaseModelAdmin(admin.ModelAdmin):

	def logiw(self, func_name, *args, type_info=True):
		if type_info: msg = '💡'
		else: msg = '⚠'
		msg += f'{self.__class__.__name__}.{func_name}'
		for arg in args: msg += f'::{arg}'
		if type_info: logging.info(msg)
		else: logging.warning(msg)

	def logw(self, func_name, *args):
		self.logiw(func_name, *args, type_info=False)

	def logi(self, func_name, *args):
		self.logiw(func_name, *args)

	def loge(self, func_name, err, *args):
		msg = f'🆘{self.__class__.__name__}.{func_name}::{err}::LINE={err.__traceback__.tb_lineno}'
		for arg in args:
			msg += f'::{arg}'
		logging.error(msg)


class ContentTypeAdmin(admin.ModelAdmin):
	list_display = ('id', 'name', 'model', 'app_label')
	list_display_links = ('name',)
	search_fields = ('id', 'model', 'app_label')#Cannot resolve keyword 'name' into field
	list_filter = ('app_label',)
admin.site.register(ContentType, ContentTypeAdmin)


class RoleAdmin(admin.ModelAdmin):
	list_display = ('id', 'weight', 'value', 'description', 'group')
	list_display_links = ('value',)
	search_fields = ('value', 'description', 'group__name')
	list_select_related = ('group',)
	list_filter = ('group',)
admin.site.register(Role, RoleAdmin)


class RoleModelAdmin(admin.ModelAdmin):
	list_display = ('id', 'value', 'description')
	list_display_links = ('value',)
	search_fields = ('value', 'description')
admin.site.register(RoleModel, RoleModelAdmin)


class RoleFieldAdmin(admin.ModelAdmin):
	list_display = ('id', 'value', 'role', 'role_model', 'read', 'write')
	list_display_links = ('value',)
	search_fields = ('value', 'role__value', 'role_model__value')
	list_select_related = ('role', 'role_model')
	list_filter = ('role', 'role_model')
admin.site.register(RoleField, RoleFieldAdmin)


class PermissionAdmin(admin.ModelAdmin):
	list_display = ('id', 'name', 'content_type', 'codename')
	list_display_links = ('name',)
	search_fields = ('id', 'name', 'codename')
admin.site.register(Permission, PermissionAdmin)


class UserAdmin(BaseUserAdmin):
	list_display = ('id', 'get_avatar', 'staff', 'get_username', 'email', 'first_name', 'last_name', 'get_role', 'last_login', 'get_groups')
	search_fields = ('username', 'email', 'first_name', 'last_name', 'last_login', 'role__value', 'role__description')
	list_select_related = ['role']
	list_filter = ('is_staff', 'is_superuser', 'is_active', 'role')
	fieldsets = (
		*BaseUserAdmin.fieldsets,#original form fieldsets, expanded
		(# new fieldset added on to the bottom
			'Custom Field Heading',#group heading of your choice; set to None for a blank space instead of a header
			{
				'fields': (
					'role',
					'avatar'
				),
			},
		),
	)
	actions = ('load_from_xls', 'set_all_push_notifications', 'selected_to_xls', 'username_lastlogin_to_xls')

	def logiw(self, func_name, *args, type_info=True):
		if type_info: msg = '💡'
		else: msg = '⚠️'
		msg += f'{self.__class__.__name__}.{func_name}'
		for arg in args: msg += f'::{arg}'
		if type_info: logging.info(msg)
		else: logging.warning(msg)

	def logw(self, *args):
		self.logiw(sys._getframe().f_back.f_code.co_name, *args, type_info=False)

	def logi(self, *args):
		self.logiw(sys._getframe().f_back.f_code.co_name, *args)

	def loge(self, err, *args):
		msg = f'🆘{self.__class__.__name__}.{err.__traceback__.tb_frame.f_code.co_name}::{err}::LINE={err.__traceback__.tb_lineno}'
		for arg in args:
			msg += f'::{arg}'
		logging.error(msg)

	def worksheet_cell_write(self, worksheet, row, col, value, type_value = None, fmt = None):
		func_write = worksheet.write
		if type_value == 'as_number':
			func_write = worksheet.write_number
		elif type_value == 'as_datetime':
			func_write = worksheet.write_datetime
		try:
			if fmt:
				func_write(row, col, value, fmt)
			else:
				func_write(row, col, value)
		except Exception as e:
			try:
				if fmt:
					func_write(row, col, repr(value), fmt)
				else:
					func_write(row, col, repr(value))
			except Exception as e:
				self.loge(e, row, col)
		return col + 1

	def worksheet_row_write(self, worksheet, row, values):
		col = 0
		for item in values:
			col = self.worksheet_cell_write(worksheet, row, col, item)
		return row + 1

	def get_queryset(self, request):
		qs = super().get_queryset(request)
		user = request.user
		if user.is_superuser:
			return qs
		return qs.filter(Q(companies__in=user.companies.all()) | Q(sale_points__in=user.sale_points.all())).exclude(role__weight__lt=user.role.weight).distinct()

	def staff(self, obj):
		return '✅' if obj.is_staff else '❌'
	staff.short_description = _('Staff')
	staff.admin_order_field = 'is_staff'

	def get_username(self, obj):
		model_meta = obj._meta
		info = (model_meta.app_label, model_meta.model_name)
		admin_url = reverse('admin:%s_%s_change' % info, args=(obj.pk,))
		return format_html('<font size="+1"><a href="{}" target="_blank">{}</a></font>', admin_url, obj.username)
	get_username.short_description = _('Login')
	get_username.admin_order_field = 'username'

	def get_role(self, obj):
		if obj.role:
			model_meta = obj.role._meta
			info = (model_meta.app_label, model_meta.model_name)
			admin_url = reverse('admin:%s_%s_change' % info, args=(obj.role_id,))
			return format_html('<p><font color="gray" face="Verdana, Geneva, sans-serif"><a href="{}" target="_blank">{}</a></font></p>', admin_url, obj.role.value)
		else:
			return ''
	get_role.short_description = _('Role')
	get_role.admin_order_field = 'role'

	def get_groups(self, obj):
		return format_html_join('\n', '<p><font color="gray" face="Verdana, Geneva, sans-serif"><a href="{}/auth/group/{}/change/" target="_blank">{}</a></font></p>', obj.groups.annotate(admin_path_prefix=Value(settings.ADMIN_PATH_PREFIX, CharField())).values_list('admin_path_prefix', 'id', 'name'))
	get_groups.short_description = _('Groups')
	get_groups.admin_order_field = 'groups'

	def get_avatar(self, obj):
		if obj.avatar:
			return format_html('<img src="{}" width="32" height="32">', obj.avatar)
		else:
			default_content = '''data:image/svg+xml,<?xml version="1.0" encoding="UTF-8" standalone="no"?><!DOCTYPE svg PUBLIC "-//W3C//DTD SVG 1.1//EN" "http://www.w3.org/Graphics/SVG/1.1/DTD/svg11.dtd"><svg version="1.1" baseProfile="full" xmlns="http://www.w3.org/2000/svg" xmlns:xlink="http://www.w3.org/1999/xlink" xmlns:ev="http://www.w3.org/2001/xml-events" width="100%" height="100%"><rect fill="white" x="0" y="0" width="100%" height="100%" /><rect fill="green" x="0" y="0" width="100%" height="100%" rx="1em"/></svg>'''
			return format_html('''<img src="{}" width="32" height="32">''', default_content)
	get_avatar.short_description = _('Avatar')

	def load_from_xls(self, request, queryset):
		form = None
		if 'apply' in request.POST:
			form = UploadFileForm(request.POST, request.FILES)
			if form.is_valid():
				file = form.cleaned_data['file']
				if file:
					if file.size > 1048576:#1MB
						self.message_user(request, f'🚫ERROR BIG FILE SIZE = {file.size} BYTES', messages.ERROR)
						return
					####################################
					output = BytesIO()
					fn = '{}.xlsx'.format(datetime.now().strftime('%Y%m%d%H%M%S'))
					workbook = xlsxwriter.Workbook(output, {'in_memory': True})
					worksheet = workbook.add_worksheet()
					row_id = 1
					####################################
					wb = load_workbook(file)
					for sheetname in wb.sheetnames:
						self.logi(sheetname)
						ws = wb[sheetname]
						list_rows = list(ws.rows)
						self.logi(list_rows[0][:10])
						for row in list_rows[1:]:
							try:
								fl,fio,rl,eml,phone = list(row)[:10]
							except Exception as e:
								self.loge(e)
							else:
								try: fl = fl.value.strip()
								except: fl = fl.value
								try: fio = fio.value.strip()
								except: fio = fio.value
								try: rl = rl.value.strip()
								except: rl = rl.value
								try: eml = eml.value.strip()
								except: eml = eml.value
								try: phone = phone.value.strip()
								except: phone = phone.value
								####################
								if not fl or len(fl) < 9 or not fio:
									continue
								else:
									sfio = fio.split(' ')
									family = sfio[0]
									name, patronymic = '', ''
									if len(sfio) > 1:
										name = sfio[1]
									if len(sfio) > 2:
										patronymic = sfio[2]
									login = slugify(family)
									if User.objects.filter(username=login).count():
										login = f'{login}{User.objects.count()}'
									password = User.objects.make_random_password()
									user_args = [login, password, eml, family, name, patronymic]
									user = None
									try:
										user = User.objects.create_user(login, eml, password, first_name=name, last_name=family)
									except Exception as e:
										self.loge(e, user_args)
										login = f'{login}{User.objects.count()}'
										try:
											user = User.objects.create_user(login, eml, password, first_name=name, last_name=family)
										except Exception as e:
											self.loge(e, user_args)
									if not user:
										self.logw('USER-NOT-CREATED', user_args)
									else:
										try:
											role = Role.objects.get(value=rl)
										except Exception as e:
											self.loge(e)
											user.role_id = 2
										else:
											user.role = role
											user.groups.set([role.group])
										#user.patronymic_text = patronymic
										try:
											user.save()
										except Exception as e:
											self.loge(e)
										else:
											self.logi(user)
										values = [login, password, eml, family, name, patronymic, rl]
										row_id = self.worksheet_row_write(worksheet, row_id, values)
					####################################
					workbook.close()
					output.seek(0)
					####################################
					self.message_user(request, fn, messages.SUCCESS)
					response = FileResponse(output, as_attachment=True, filename=fn)
					return response
				else:
					self.message_user(request, f'🚫ERROR LOAD FILE', messages.ERROR)
					return
			else:
				self.message_user(request, f'🚫ERROR WITH FORM SELECT FILE', messages.ERROR)
				return
		if not form:
			form = UploadFileForm(initial={'_selected_action': request.POST.getlist(admin.helpers.ACTION_CHECKBOX_NAME)})
		m = queryset.model._meta
		context = {}
		context['items'] = queryset
		context['form'] = form
		context['title'] = _('File')
		context['current_action'] = sys._getframe().f_code.co_name
		context['subtitle'] = 'admin_select_file_form'
		context['site_title'] = _('Users')
		context['is_popup'] = True
		context['is_nav_sidebar_enabled'] = True
		context['site_header'] = _('Support')
		context['has_permission'] = True
		context['site_url'] = reverse('admin:{}_{}_changelist'.format(m.app_label, m.model_name))
		context['available_apps'] = (m.app_label,)
		context['app_label'] = m.app_label
		return render(request, 'admin_select_file_form.html', context)
	load_from_xls.short_description = f'⚔️{_("load from XLSX file")}'

	def set_all_push_notifications(self, request, queryset):
		changed = 0
		try:
			push_type = get_model('core.NotificationType').objects.get(value='push')
		except Exception as e:
			self.loge(e)
			self.message_user(request, f'{e}', messages.ERROR)
			return
		sources = get_model('core.NotificationSource').objects.all()
		self.logi(f'Notification Sources = {sources.count()}')
		for user in queryset:
			self.logi(f'USER = {user}')
			for src in sources:
				ntfopt = None
				try:
					ntfopt = get_model('core.NotificationOption').objects.get(owner_id=user.id, source_id=src.id)
				except ObjectDoesNotExist as e:
					try:
						ntfopt = get_model('core.NotificationOption')(owner_id=user.id, source_id=src.id)
					except Exception as e:
						self.loge(e)
					else:
						try:
							ntfopt.save()
						except Exception as e:
							self.loge(e)
				except Exception as e:
					self.loge(e)
				if ntfopt:
					ntfopt.types.add(push_type)
					changed += 1
		self.message_user(request, f'🆗 {changed}', messages.SUCCESS)
	set_all_push_notifications.short_description = f'🔔{_("set all push notifications")}'

	def queryset_to_xls(self, request, queryset, field_names=[], exclude_fields=['password','favorites','notificationoption','notificationdelay','notificationtask', 'avatar', 'groups', 'user_permissions']):
		output = None
		if queryset.count():
			if not field_names:
				for field in queryset.model._meta.get_fields():
					if field.name and field.name not in exclude_fields:
						field_names.append(field.name)
			output = BytesIO()
			workbook = xlsxwriter.Workbook(output, {'in_memory': True})
			worksheet = workbook.add_worksheet()
			#date_format = workbook.add_format({'num_format': 'mmmm d yyyy'})
			#money_format = workbook.add_format({'num_format': '$#,##0'})
			col = 0
			for field_name in field_names:
				col = self.worksheet_cell_write(worksheet, 0, col, field_name)
			row = 1
			for item in queryset:
				col = 0
				for field_name in field_names:
					if not hasattr(item, field_name):
						col += 1
						continue
					else:
						if not field_name:
							col += 1
							continue
					try:
						value = getattr(item, field_name)
					except AttributeError as e:
						col += 1
						self.loge(e)
					except Exception as e:
						col += 1
						self.loge(e)
					else:
						if not value:
							col += 1
						else:
							format_value = None
							tvalue = None
							if isinstance(value, datetime):
								#tvalue = 'as_datetime'
								#format_value = date_format
								value = f'{value.strftime("%Y.%m.%d %H:%M:%S")}'
							elif isinstance(value, (int, float)):#value.isdigit():
								#format_value = money_format
								tvalue = 'as_number'
							elif not isinstance(value, str):
								value = f'{value}'
							col = self.worksheet_cell_write(worksheet, row, col, value, tvalue, format_value)
				row += 1
			workbook.close()
			output.seek(0)
		return output

	def selected_to_xls(self, request, queryset):
		output = self.queryset_to_xls(request, queryset)
		if output:
			fn = '{}.xlsx'.format(timezone.now().strftime('%Y%m%d%H%M%S'))
			self.message_user(request, f'🆗 {_("Finished")} ✏️({fn})', messages.SUCCESS)
			return FileResponse(output, as_attachment=True, filename=fn)
		self.message_user(request, _('please select items'), messages.ERROR)
	selected_to_xls.short_description = f'🏗️{_("selected to xls file")}'

	def username_lastlogin_to_xls(self, request, queryset):
		output = self.queryset_to_xls(request, queryset, ['username', 'first_name', 'last_name', 'last_login'])
		if output:
			fn = '{}.xlsx'.format(timezone.now().strftime('%Y%m%d%H%M%S'))
			self.message_user(request, f'🆗 {_("Finished")} ✏️({fn})', messages.SUCCESS)
			return FileResponse(output, as_attachment=True, filename=fn)
		self.message_user(request, _('please select items'), messages.ERROR)
	username_lastlogin_to_xls.short_description = f'🏗️{_("username+last_login to xls file")}'

admin.site.register(User, UserAdmin)
#admin.site.unregister(Group)
